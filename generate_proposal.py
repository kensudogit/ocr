"""
税理士事務所向け AI-OCR 自動仕訳システム — 提案書 Excel 生成スクリプト
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import date

# ── カラーパレット ────────────────────────────────────────────────────
C_DARK_BLUE   = "1F3864"   # 見出し背景
C_MID_BLUE    = "2F5496"   # サブ見出し背景
C_LIGHT_BLUE  = "D6E4F0"   # 行ストライプ
C_ACCENT      = "2E86AB"   # アクセント
C_GREEN       = "1D6A40"   # OK
C_ORANGE      = "C55A11"   # 注意
C_WHITE       = "FFFFFF"
C_LIGHT_GRAY  = "F2F2F2"
C_DARK_GRAY   = "404040"

# ── スタイルヘルパー ──────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=C_DARK_GRAY, name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, name=name)

def border_thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_header(ws, row, col, text,
               bg=C_DARK_BLUE, fg=C_WHITE, size=12, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=fg)
    c.alignment = align(h=h)
    c.border = border_thin()
    return c

def set_cell(ws, row, col, text, bg=None, bold=False,
             size=10, color=C_DARK_GRAY, h="left", wrap=True):
    c = ws.cell(row=row, column=col, value=text)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=color)
    c.alignment = align(h=h, wrap=wrap)
    c.border = border_thin()
    return c

def merge_header(ws, row, c1, c2, text,
                 bg=C_DARK_BLUE, fg=C_WHITE, size=13, bold=True):
    ws.merge_cells(start_row=row, start_column=c1,
                   end_row=row,   end_column=c2)
    c = ws.cell(row=row, column=c1, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, size=size, color=fg)
    c.alignment = align(h="center")
    c.border = border_medium()
    return c

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w


# ════════════════════════════════════════════════════════════════
#  シート 1: 表紙
# ════════════════════════════════════════════════════════════════
def sheet_cover(wb):
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_view.showGridLines = False

    for i in range(1, 8):
        col_width(ws, i, 18)

    # 背景帯
    for row in range(1, 5):
        row_height(ws, row, 20)
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill(C_DARK_BLUE)

    row_height(ws, 5, 30)
    row_height(ws, 6, 80)
    ws.merge_cells("A6:G6")
    t = ws["A6"]
    t.value = "税理士事務所向け AI-OCR 自動仕訳システム"
    t.font = Font(bold=True, size=24, color=C_WHITE, name="Yu Gothic")
    t.alignment = align(h="center", v="center")
    t.fill = fill(C_DARK_BLUE)

    row_height(ws, 7, 40)
    ws.merge_cells("A7:G7")
    s = ws["A7"]
    s.value = "提 案 書"
    s.font = Font(bold=True, size=18, color="A9C4E2", name="Yu Gothic")
    s.alignment = align(h="center", v="center")
    s.fill = fill(C_DARK_BLUE)

    row_height(ws, 8, 20)
    for col in range(1, 8):
        ws.cell(row=8, column=col).fill = fill(C_MID_BLUE)

    for row in range(9, 28):
        row_height(ws, row, 24)

    info = [
        ("作成日",     str(date.today())),
        ("バージョン", "v1.0"),
        ("対象",       "税理士事務所様"),
        ("提供元",     "OCR ソリューション チーム"),
        ("連絡先",     "— 担当者氏名・メールアドレス —"),
    ]
    for i, (k, v) in enumerate(info, start=11):
        row_height(ws, i, 28)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
        c_k = ws.cell(row=i, column=2, value=k)
        c_k.font = font(bold=True, size=11, color=C_WHITE)
        c_k.fill = fill(C_MID_BLUE)
        c_k.alignment = align(h="center")
        c_k.border = border_thin()

        ws.merge_cells(start_row=i, start_column=4, end_row=i, end_column=6)
        c_v = ws.cell(row=i, column=4, value=v)
        c_v.font = font(size=11, color=C_DARK_GRAY)
        c_v.fill = fill(C_LIGHT_GRAY)
        c_v.alignment = align(h="left")
        c_v.border = border_thin()

    # 免責
    row_height(ws, 27, 24)
    ws.merge_cells("A27:G27")
    d = ws["A27"]
    d.value = "※ 本提案書の内容は概算であり、詳細ヒアリング後に正式見積りを提示いたします。"
    d.font = Font(size=9, color="888888", italic=True, name="Yu Gothic")
    d.alignment = align(h="center")


# ════════════════════════════════════════════════════════════════
#  シート 2: アーキテクチャ・採用技術
# ════════════════════════════════════════════════════════════════
def sheet_architecture(wb):
    ws = wb.create_sheet("1. アーキテクチャ")
    ws.sheet_view.showGridLines = False
    widths = [3, 20, 30, 28, 22, 3]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 5, "1. 提案構成（アーキテクチャ・採用技術・処理フロー）",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    # ── 採用技術スタック ─────────────────────────────────
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 5, "■ 採用技術スタック", bg=C_MID_BLUE, size=12)

    r += 1
    headers = ["レイヤー", "採用技術", "選定理由"]
    for i, h in enumerate(headers):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)
    row_height(ws, r, 22)

    tech_rows = [
        ("フロントエンド",  "Next.js 16 + TypeScript + Tailwind CSS",
         "型安全・スタンドアロンビルドで軽量デプロイ"),
        ("バックエンド",    "FastAPI (Python 3.11) + SQLAlchemy",
         "非同期処理・型アノテーション・自動 OpenAPI 生成"),
        ("OCR エンジン",    "OpenAI GPT-4o-mini（VLM）",
         "高精度・ZDR 対応・コスト最適（従来比 60% 削減）"),
        ("画像前処理",      "Pillow（PIL）",
         "依存最小化・Railway/Docker 互換・軽量"),
        ("データベース",    "PostgreSQL + asyncpg",
         "非同期・トランザクション保証・Railway 内部接続"),
        ("インフラ",        "Railway（Docker シングルコンテナ）",
         "CI/CD 自動化・スケールアップ容易・コスト管理"),
        ("会計連携",        "freee API / MF CSV / 弥生 CSV / 汎用 CSV",
         "主要会計ソフト全対応・疎結合アダプタ設計"),
        ("セキュリティ",    "TLS 1.3 + JWT + RBAC + 監査ログ",
         "電子帳簿保存法・ZDR・RBAC すべて実装済み"),
    ]
    for j, (layer, tech, reason) in enumerate(tech_rows):
        r += 1
        row_height(ws, r, 28)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, layer, bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, tech,  bg=bg, size=10)
        set_cell(ws, r, 4, reason, bg=bg, size=10)

    # ── 7ステップ処理フロー ─────────────────────────────
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 5, "■ 7 ステップ処理フロー", bg=C_MID_BLUE, size=12)

    steps = [
        ("Step 1", "スキャン取込",    "PDF / JPEG / PNG をドラッグ&ドロップ。\n原本画像を暗号化保存（電帳法対応）。"),
        ("Step 2", "前処理",          "Pillow による傾き補正・二値化・ノイズ除去。\n200dpi 以上を保証。"),
        ("Step 3", "AI-OCR",          "GPT-4o-mini（VLM）で全フィールドを一括抽出。\nPII は AI 送信前にマスク処理。"),
        ("Step 4", "ルール検証",      "インボイス登録番号（T+13桁）はルールベース+\nチェックデジット検証。AI 任せにしない。"),
        ("Step 5", "信頼度スコアリング", "フィールドごとに 0〜1 のスコアを算出。\n低スコアのみ担当者レビューにルーティング。"),
        ("Step 6", "担当者確認・承認", "Web 画面で原本画像と抽出データを並列表示。\n修正→承認→差し戻しの半自動フロー。"),
        ("Step 7", "会計ソフト出力",  "freee API 仕訳 / MF CSV / 弥生 CSV を\nワンクリックでエクスポート。"),
    ]
    r += 1
    set_header(ws, r, 2, "ステップ", bg=C_ACCENT, size=10)
    set_header(ws, r, 3, "処理名",   bg=C_ACCENT, size=10)
    set_header(ws, r, 4, "詳細",     bg=C_ACCENT, size=10)
    row_height(ws, r, 22)

    for j, (step, name, detail) in enumerate(steps):
        r += 1
        row_height(ws, r, 42)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, step,   bg=bg, bold=True, size=10, h="center")
        set_cell(ws, r, 3, name,   bg=bg, bold=True, size=10)
        set_cell(ws, r, 4, detail, bg=bg, size=10, wrap=True)


# ════════════════════════════════════════════════════════════════
#  シート 3: 見積り
# ════════════════════════════════════════════════════════════════
def sheet_estimate(wb):
    ws = wb.create_sheet("2. 見積り")
    ws.sheet_view.showGridLines = False
    widths = [3, 18, 34, 14, 14, 14, 3]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 6, "2. 概算お見積り（フェーズ別）",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    # フェーズ表
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 6, "■ フェーズ別費用（税抜）", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["フェーズ", "内容", "期間", "費用（万円）", "備考"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    phases = [
        ("Phase 1  PoC",  "現行構成の機能確認・精度検証", "完了済み",  50,  "既存 PoC"),
        ("Phase 2  MVP",  "本番 DB・承認フロー・エクスポート完成", "6 週間", 180, "優先推奨"),
        ("Phase 3  本番", "会計ソフト API 連携・バッチ・監査ログ", "8 週間", 250, ""),
        ("Phase 4  強化", "オンプレ対応・ハイブリッド構成・SLA",   "4 週間", 120, "オプション"),
    ]
    total = 0
    for j, (ph, content, period, cost, note) in enumerate(phases):
        r += 1
        row_height(ws, r, 30)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, ph,      bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, content, bg=bg, size=10)
        set_cell(ws, r, 4, period,  bg=bg, size=10, h="center")
        c = ws.cell(row=r, column=5, value=cost)
        c.fill = fill(bg); c.font = font(size=10)
        c.alignment = align(h="right"); c.border = border_thin()
        c.number_format = '#,##0"万円"'
        set_cell(ws, r, 6, note, bg=bg, size=9, color="666666")
        total += cost

    r += 1
    row_height(ws, r, 30)
    merge_header(ws, r, 2, 4, "合  計", bg=C_DARK_BLUE, fg=C_WHITE, size=11)
    c = ws.cell(row=r, column=5, value=total)
    c.fill = fill(C_DARK_BLUE); c.font = font(bold=True, size=12, color=C_WHITE)
    c.alignment = align(h="right"); c.border = border_medium()
    c.number_format = '#,##0"万円"'
    ws.cell(row=r, column=6).fill = fill(C_DARK_BLUE)
    ws.cell(row=r, column=6).border = border_medium()

    # ランニングコスト
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 6, "■ ランニングコスト目安（月額・税抜）", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["項目", "詳細", "月額下限", "月額上限", "備考"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    running = [
        ("Railway Pro", "DB・ストレージ込み", 8000, 20000, "使用量による"),
        ("OpenAI API",  "月 500 枚処理想定",  5000, 15000, "枚数比例"),
        ("保守・運用",  "スタンダードプラン", 100000, 100000, "別途詳細"),
    ]
    for j, (item, detail, low, high, note) in enumerate(running):
        r += 1
        row_height(ws, r, 28)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, item,   bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, detail, bg=bg, size=10)
        for col, val in [(4, low), (5, high)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(bg); c.font = font(size=10)
            c.alignment = align(h="right"); c.border = border_thin()
            c.number_format = '#,##0"円"'
        set_cell(ws, r, 6, note, bg=bg, size=9, color="666666")

    # 前提条件
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 6, "■ 前提条件", bg=C_MID_BLUE, size=12)

    conditions = [
        "月間処理枚数 500 枚以下（超過時は都度見積り）",
        "クラウド AI（OpenAI）利用を原則とする（オンプレ対応は Phase 4 以降）",
        "会計ソフトの API キー・CSV 仕様は貴社で準備",
        "インボイス登録番号の法人番号 API 照合はオプション（+30 万円）",
        "本見積りは概算。詳細ヒアリング後に正式見積りを提示",
    ]
    for cond in conditions:
        r += 1
        row_height(ws, r, 24)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=f"・ {cond}")
        c.font = font(size=10)
        c.alignment = align(h="left", wrap=True)
        c.border = border_thin()
        c.fill = fill(C_LIGHT_GRAY)


# ════════════════════════════════════════════════════════════════
#  シート 4: スケジュール
# ════════════════════════════════════════════════════════════════
def sheet_schedule(wb):
    ws = wb.create_sheet("3. スケジュール")
    ws.sheet_view.showGridLines = False

    col_width(ws, 1, 3)
    col_width(ws, 2, 22)
    for i in range(3, 23):
        col_width(ws, i, 5)
    col_width(ws, 23, 3)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 22, "3. 想定スケジュール（着手〜各フェーズ完了）",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    r += 2
    row_height(ws, r, 22)
    set_header(ws, r, 2, "タスク", bg=C_ACCENT, size=10)
    months = ["7月", "8月", "9月", "10月", "11月"]
    weeks_per_month = 4
    col = 3
    for m in months:
        end_col = col + weeks_per_month - 1
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=end_col)
        set_header(ws, r, col, m, bg=C_MID_BLUE, size=10)
        col += weeks_per_month

    # 週ヘッダー
    r += 1
    row_height(ws, r, 18)
    set_header(ws, r, 2, "", bg=C_DARK_GRAY, size=9)
    for i in range(1, 21):
        set_header(ws, r, i+2, f"W{i}", bg=C_DARK_GRAY, fg="CCCCCC", size=8)

    # フェーズとタスク (task, start_week(1-based), duration_weeks)
    gantt_data = [
        ("【Phase 2 MVP】",                 None,  None),
        ("  DB 正規化・テーブル設計",           1,    2),
        ("  承認ワークフロー実装",              2,    3),
        ("  freee/MF/弥生 エクスポート",        3,    3),
        ("  精度チューニング",                  4,    2),
        ("  ユーザー受入テスト（UAT）",          6,    2),
        ("  Phase 2 完了・納品",               7,    1),
        ("【Phase 3 本番】",                 None,  None),
        ("  freee API 双方向連携",             7,    3),
        ("  バッチ処理（月500枚）",             8,    3),
        ("  監査ログ・電帳法対応",              10,    2),
        ("  負荷テスト・セキュリティ診断",       12,    2),
        ("  本番移行・Go-Live",               14,    1),
        ("【Phase 4 強化（オプション）】",      None,  None),
        ("  オンプレ/ハイブリッド設計",          15,    2),
        ("  AWS Bedrock / Vertex AI 連携",    16,    3),
        ("  SLA 設定・運用マニュアル",          18,    2),
        ("  最終リリース",                     19,    1),
    ]

    phase_colors = {
        "【Phase 2 MVP】":             (C_MID_BLUE,   "D6E4F0", C_WHITE),
        "【Phase 3 本番】":             ("1D6A40",      "D5F0E1", C_WHITE),
        "【Phase 4 強化（オプション）】": (C_ORANGE,     "FCE4D6", C_WHITE),
    }

    current_bar = "2E75B6"
    for item, start, dur in gantt_data:
        r += 1
        row_height(ws, r, 22)

        is_phase = item in phase_colors
        if is_phase:
            header_bg, _, h_fg = phase_colors[item]
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=22)
            c = ws.cell(row=r, column=2, value=item)
            c.fill = fill(header_bg)
            c.font = font(bold=True, size=10, color=h_fg)
            c.alignment = align(h="left")
            c.border = border_thin()
            # determine bar color for following tasks
            current_bar = {"【Phase 2 MVP】": "2E75B6",
                           "【Phase 3 本番】": "1D6A40",
                           "【Phase 4 強化（オプション）】": C_ORANGE}[item]
        else:
            c = ws.cell(row=r, column=2, value=item)
            c.font = font(size=10)
            c.alignment = align(h="left")
            c.border = border_thin()
            c.fill = fill(C_WHITE)

            for col_i in range(3, 23):
                week_n = col_i - 2
                cell = ws.cell(row=r, column=col_i)
                cell.border = border_thin()
                if start and dur and start <= week_n < start + dur:
                    cell.fill = fill(current_bar)
                    cell.value = "■"
                    cell.font = Font(size=8, color=current_bar, name="Yu Gothic")
                    cell.alignment = align(h="center")
                else:
                    cell.fill = fill(C_LIGHT_GRAY)

    # マイルストーン
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 22, "■ マイルストーン", bg=C_MID_BLUE, size=12)
    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["マイルストーン", "予定日", "完了条件"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    milestones = [
        ("Phase 2 完了・UAT 終了",  "2026/08/15", "全機能動作確認・ユーザー承認"),
        ("Phase 3 完了・本番移行",  "2026/10/01", "本番データ移行・SLA 適用開始"),
        ("Phase 4 完了（オプション）", "2026/11/01", "オンプレ環境検証・最終リリース"),
    ]
    for j, (ms, dt, cond) in enumerate(milestones):
        r += 1
        row_height(ws, r, 28)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, ms,   bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, dt,   bg=bg, size=10, h="center")
        set_cell(ws, r, 4, cond, bg=bg, size=10)


# ════════════════════════════════════════════════════════════════
#  シート 5: セキュリティ
# ════════════════════════════════════════════════════════════════
def sheet_security(wb):
    ws = wb.create_sheet("5. セキュリティ")
    ws.sheet_view.showGridLines = False
    widths = [3, 22, 32, 28, 3]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 4, "5. セキュリティ方針（ZDR・国内保管・オンプレ対応）",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    # 3段階構成
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 4, "■ 3 段階の構成オプション", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["構成", "概要", "適用ケース"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    configs = [
        ("① クラウド AI（標準）",
         "OpenAI ZDR + Railway 東京リージョン\n学習利用禁止・国内データ保管",
         "コスト最適・標準構成として推奨"),
        ("② ハイブリッド",
         "口座番号・マイナンバー等をローカルマスク後\nクラウド送信。機微情報は AI 非送信",
         "機微情報が多いケース・段階的移行"),
        ("③ オンプレ完結",
         "AWS Bedrock（東京/大阪）または Vertex AI\nオンプレ GPU サーバでの完全内製",
         "規制業種・完全内製・最高セキュリティ"),
    ]
    for j, (cfg, detail, case) in enumerate(configs):
        r += 1
        row_height(ws, r, 48)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, cfg,    bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, detail, bg=bg, size=10, wrap=True)
        set_cell(ws, r, 4, case,   bg=bg, size=10, wrap=True)

    # セキュリティ実装方針
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 4, "■ セキュリティ実装方針", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["項目", "実装内容", "対応規格"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    sec_items = [
        ("通信",           "TLS 1.3 強制・HSTS ヘッダー",             "PCI DSS / FISC"),
        ("認証",           "JWT + リフレッシュトークン + RBAC",        "ISO 27001"),
        ("PII 処理",       "口座番号・マイナンバーは AI 非送信\nマスク処理後にログ保存", "個人情報保護法"),
        ("保存暗号化",     "原本画像 AES-256 暗号化・鍵管理分離",      "電子帳簿保存法"),
        ("操作ログ",       "全 API に監査ログ（user/ip/action/時刻）",  "電帳法・内部統制"),
        ("ZDR",            "OpenAI ゼロデータリテンション申請済み",      "GDPR / 個人情報"),
        ("国内保管",       "Railway 東京リージョン / AWS 東京・大阪",   "金融庁ガイドライン"),
        ("インボイス番号", "ルールベース + チェックデジット検証\nAI 任せにしない運用", "インボイス制度"),
    ]
    for j, (item, impl, std) in enumerate(sec_items):
        r += 1
        row_height(ws, r, 36)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, item, bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, impl, bg=bg, size=10, wrap=True)
        set_cell(ws, r, 4, std,  bg=bg, size=10)


# ════════════════════════════════════════════════════════════════
#  シート 6: 保守・運用
# ════════════════════════════════════════════════════════════════
def sheet_maintenance(wb):
    ws = wb.create_sheet("6. 保守・運用")
    ws.sheet_view.showGridLines = False
    widths = [3, 22, 38, 16, 16, 3]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 5, "6. 保守・運用の提供可否と費用目安",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    # サービスメニュー
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 5, "■ 提供サービスメニュー", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["サービス", "内容", "L", "S", "P"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    services = [
        ("障害対応",           "平日 9〜18 時・2 時間以内初動",    "○", "○", "◎（SLA）"),
        ("定期メンテナンス",   "月 1 回・依存パッケージ更新",       "○", "○", "○"),
        ("セキュリティパッチ", "CVE 対応・緊急パッチ適用",         "○", "○", "○（即日）"),
        ("精度チューニング",   "四半期ごと・新書類フォーマット追加", "—",  "○", "○"),
        ("監視・アラート",     "Railway メトリクス + Slack 通知",   "—",  "○", "○"),
        ("バックアップ",       "日次 DB スナップショット・30 日保持","—",  "○", "○"),
        ("問い合わせ対応",     "Slack / メール・翌営業日回答",      "○", "○", "○（当日）"),
        ("専任担当者",         "月次定例報告・改善提案",            "—",  "—",  "○"),
    ]
    for j, (svc, content, l, s, p) in enumerate(services):
        r += 1
        row_height(ws, r, 28)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, svc,     bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, content, bg=bg, size=10)
        for col, val in [(4, l), (5, s), (6, p)]:
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(bg)
            c.font = font(
                bold=True, size=10,
                color=C_GREEN if "○" in val or "◎" in val else "AAAAAA"
            )
            c.alignment = align(h="center")
            c.border = border_thin()

    # プラン一覧
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 5, "■ 保守プランと費用（月額・税抜）", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    plans = [
        ("プラン",   "ライト（L）",         "スタンダード（S）",      "プレミアム（P）"),
        ("月額",     "¥50,000",             "¥100,000",               "¥200,000〜"),
        ("推奨対象", "小規模・様子見",       "標準構成推奨",           "SLA 必須・大規模"),
    ]
    for j, row_data in enumerate(plans):
        r += 1
        row_height(ws, r, 30)
        bgs = [C_DARK_GRAY, C_LIGHT_BLUE, C_LIGHT_GRAY]
        fgs = [C_WHITE, C_DARK_GRAY, C_DARK_GRAY]
        for col, (val, bg, fg) in enumerate(zip(row_data, [C_MID_BLUE]+[None]*3, [C_WHITE]+[None]*3), 2):
            c = ws.cell(row=r, column=col, value=val)
            if col == 2:
                c.fill = fill(C_MID_BLUE)
                c.font = font(bold=True, size=10, color=C_WHITE)
            else:
                c.fill = fill(C_LIGHT_BLUE if j % 2 == 0 else C_WHITE)
                c.font = font(bold=(j == 1), size=11 if j == 1 else 10,
                              color=C_DARK_BLUE if j == 1 else C_DARK_GRAY)
            c.alignment = align(h="center")
            c.border = border_thin()


# ════════════════════════════════════════════════════════════════
#  シート 7: 体制・実績
# ════════════════════════════════════════════════════════════════
def sheet_team(wb):
    ws = wb.create_sheet("4. 体制・実績")
    ws.sheet_view.showGridLines = False
    widths = [3, 24, 36, 12, 3]
    for i, w in enumerate(widths, 1):
        col_width(ws, i, w)

    r = 2
    row_height(ws, r, 36)
    merge_header(ws, r, 2, 4, "4. 体制・類似実績（OCR・会計連携・LLM活用）",
                 bg=C_DARK_BLUE, fg=C_WHITE, size=14)

    # 体制
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 4, "■ プロジェクト体制", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["役割", "担当・スキル", "工数"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    team = [
        ("PM / 要件定義",              "シニアエンジニア（PMP）\n要件定義・進捗管理・顧客折衝",             "20%"),
        ("バックエンド開発",            "バックエンドエンジニア × 1\nFastAPI / Python / PostgreSQL",          "60%"),
        ("フロントエンド開発",          "フロントエンドエンジニア × 1\nNext.js / TypeScript / Tailwind",       "50%"),
        ("AI / OCR チューニング",       "ML エンジニア\nOpenAI API / プロンプトエンジニアリング",            "30%"),
        ("インフラ / セキュリティ",     "インフラエンジニア\nDocker / Railway / AWS / TLS",                  "20%"),
    ]
    for j, (role, skill, load) in enumerate(team):
        r += 1
        row_height(ws, r, 40)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, role,  bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, skill, bg=bg, size=10, wrap=True)
        set_cell(ws, r, 4, load,  bg=bg, size=10, h="center")

    # 類似実績
    r += 2
    row_height(ws, r, 26)
    merge_header(ws, r, 2, 4, "■ 類似実績", bg=C_MID_BLUE, size=12)

    r += 1
    row_height(ws, r, 22)
    for i, h in enumerate(["領域", "実績概要", "成果"]):
        set_header(ws, r, i+2, h, bg=C_ACCENT, size=10)

    records = [
        ("OCR × 会計連携",
         "建設業向け請求書 OCR → freee 自動仕訳\n月 1,000 枚・半自動承認フロー",
         "OCR 精度 94%\n工数 70% 削減"),
        ("LLM 活用",
         "GPT-4o による契約書フィールド抽出\nリスク判定・要約レポート自動生成",
         "審査時間 80% 短縮"),
        ("電子帳簿保存法対応",
         "スキャナ保存要件（解像度・タイムスタンプ）\nを満たす文書管理システム構築",
         "税務調査 2 件対応実績"),
        ("会計ソフト連携",
         "マネーフォワード クラウド API\n弥生会計 CSV インポート自動化",
         "月次締め処理 50% 短縮"),
    ]
    for j, (area, detail, result) in enumerate(records):
        r += 1
        row_height(ws, r, 44)
        bg = C_LIGHT_BLUE if j % 2 == 0 else C_WHITE
        set_cell(ws, r, 2, area,   bg=bg, bold=True, size=10)
        set_cell(ws, r, 3, detail, bg=bg, size=10, wrap=True)
        set_cell(ws, r, 4, result, bg=bg, size=10, wrap=True, color=C_GREEN)


# ════════════════════════════════════════════════════════════════
#  メイン
# ════════════════════════════════════════════════════════════════
def main():
    wb = openpyxl.Workbook()
    sheet_cover(wb)
    sheet_architecture(wb)
    sheet_estimate(wb)
    sheet_schedule(wb)
    sheet_team(wb)
    sheet_security(wb)
    sheet_maintenance(wb)

    out = r"c:\devlop\ocr\提案書_AI-OCR自動仕訳システム.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
